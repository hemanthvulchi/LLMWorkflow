
from PySide6 import QtWidgets
from PySide6.QtWidgets import QLabel, QTextEdit, QLineEdit, QPushButton, QComboBox, QFileDialog, QMessageBox, QVBoxLayout, QFormLayout
from core.node import Node
from core.configdialog import ConfigDialog
from core.openaiconnection import OpenAIConnection
from openpyxl.utils import range_boundaries
from util.display import Display
import openpyxl
import pandas as pd
import json
import sys

class ExcelProcess2_Node(Node):
    def __init__(self):
        super().__init__()

        self.title_text = "Excel Processor"
        self.type_text = "Select a range to process data"
        self.set_color(title_color=(255, 165, 0))
        self.pin_output = self.add_pin(name="value", is_output=True)
        self.build()
        self.con = OpenAIConnection()
        self.openAIclient = self.con.get_connection()

        self.config = {
            "user_prompt": "",
            "system_prompt": "You will be provided with a message, and your task is to write a detailed report.",
            "output":"",
            "max_tokens": 64,
            "model": "gpt-3.5-turbo",
            "id": "",
            "object": "",
            "temperature": ""
        }

        # Create a single instance of the ExtendedConfigDialog
        self.dialog = ExtendedConfigDialog(json.dumps(self.config), self.openAIclient, self.widget)
        
    def init_widget(self):
        self.widget = QtWidgets.QWidget()
        self.widget.setFixedWidth(150)
        layout = QtWidgets.QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        self.textbox = QTextEdit()
        self.textbox.setFixedHeight(50)
        self.responsetext = QTextEdit()
        self.responsetext.setFixedHeight(50)
        self.responsetext.setStyleSheet("""
        QTextEdit{
        background: rgb(50, 50, 50); /*background color */
        }
        """)
        self.responsetext.setReadOnly(True)

        self.btn = QtWidgets.QPushButton("Run")
        self.btn.clicked.connect(self.btn_chat)

        self.config_btn = QtWidgets.QPushButton("Configure")
        self.config_btn.clicked.connect(self.show_configuration)

        layout.addWidget(self.textbox)
        layout.addWidget(self.responsetext)
        layout.addWidget(self.btn)
        layout.addWidget(self.config_btn)
        self.widget.setLayout(layout)

        proxy = QtWidgets.QGraphicsProxyWidget()
        proxy.setWidget(self.widget)
        proxy.setParentItem(self)

        super().init_widget()

    def show_configuration(self):
        if self.dialog.exec():
            self.config = json.loads(self.dialog.get_configuration())
            self.responsetext.setText(self.config["output"])
            self.textbox.setText(self.config["user_prompt"])
            self.pin_output.set_data(self.responsetext.toPlainText())
        print("Pin output data:", self.pin_output.get_data())

    def btn_chat(self):
        response = self.openAIclient.chat.completions.create(
            model=self.config["model"],
            messages=[
                {
                    "role": "system",
                    "content": self.config["system_prompt"]
                },
                {
                    "role": "user",
                    "content": self.textbox.toPlainText()
                }
            ],
            temperature=0.8,
            max_tokens=self.config["max_tokens"],
            top_p=1
        )
        self.responsetext.setText(response.choices[0].message.content)
        self.pin_output.set_data(self.responsetext.toPlainText())
        print("Pin output data:", self.pin_output.get_data())


class ExtendedConfigDialog(ConfigDialog):

    
    def __init__(self, config_json, openAIclient, parent=None):
        super(ExtendedConfigDialog, self).__init__(config_json, openAIclient, parent)
        self.min_col = 1
        self.min_row = 1 
        self.max_col = 1
        self.max_row = 1
        self.extra_field1 = QLineEdit()
        self.extra_field2 = QLineEdit()

        self.extra_field1.setPlaceholderText("Enter additional information 1")
        self.extra_field2.setPlaceholderText("Enter additional information 2")

        self.layout.addRow("Extra Field 1:", self.extra_field1)
        self.layout.addRow("Extra Field 2:", self.extra_field2)

        self.file_label = QLabel('Selected file:')
        self.file_label = QLabel('Selected file:')
        
        self.sheet_label = QLabel('Sheet name:')

        self.range_input = QLineEdit(self)
        self.range_input.setPlaceholderText('Enter Excel range (e.g., A1:B10)')

        self.sheet_combobox = QComboBox(self)
        self.sheet_combobox.setEnabled(False)

        self.load_button = QPushButton('Load Excel File', self)
        self.load_button.clicked.connect(self.load_file)

        self.get_data_button = QPushButton('Get Data', self)
        self.get_data_button.clicked.connect(self.get_range)
        self.get_data_button.setEnabled(False)

        self.save_data_button = QPushButton('Save Data', self)
        self.save_data_button.clicked.connect(self.save_data)
        self.save_data_button.setEnabled(False)

        self.top_layout = QFormLayout()
        self.top_layout.addWidget(self.load_button)
        self.top_layout.addWidget(self.sheet_label)
        self.top_layout.addWidget(self.file_label)
        self.top_layout.addWidget(self.range_input)
        self.top_layout.addWidget(self.sheet_combobox)
        self.top_layout.addWidget(self.get_data_button)
        self.top_layout.addWidget(self.save_data_button)
        
        self.main_layout.addLayout(self.top_layout)
        self.setmainlayout()
        self.setLayout(self.main_layout)

        # Initialize message box attributes
        self.message_box = None

    def get_configuration(self):
        config = json.loads(super().get_configuration())
        config["extra_field1"] = self.extra_field1.text()
        config["extra_field2"] = self.extra_field2.text()
        return json.dumps(config)

    def load_file(self):
        file_dialog = QFileDialog()
        file_dialog.setFileMode(QFileDialog.ExistingFile)
        file_dialog.setNameFilter("Excel Files (*.xlsx *.xls)")

        if file_dialog.exec():
            self.selected_file = file_dialog.selectedFiles()[0]
            self.file_label.setText(f'Selected file: {self.selected_file}')

            xls = pd.ExcelFile(self.selected_file)
            self.sheet_combobox.clear()
            self.sheet_combobox.addItems(xls.sheet_names)
            if xls.sheet_names:
                self.sheet_combobox.setCurrentIndex(0)
                self.sheet_combobox.setEnabled(True)
                self.get_data_button.setEnabled(True)
                self.save_data_button.setEnabled(True)
            
    def get_range(self):
        if not self.selected_file:
            Display.show_message_box('Error', 'No file selected.')
            Display.show_message_box("error","sdfs")
            return

        range_str = self.range_input.text()
        self.sheet_name = self.sheet_combobox.currentText()

        try:
            df = pd.read_excel(self.selected_file, sheet_name=self.sheet_name, engine='openpyxl')
            if range_str:
                df = self.get_data_in_range(df, range_str)
            self.data = df
            Display.show_message_box('Success', f'Selected range:\n{df}')
        except Exception as e:
            self.show_message_box('Error', str(e))

    def get_data_in_range(self, df, range_str):
        self.min_col, self.min_row, self.max_col, self.max_row = range_boundaries(range_str)
        #self.min_col -= 1
        #self.max_col -= 1
        return df.iloc[self.min_row-1:self.max_row, self.min_col:self.max_col]

    def save_data(self):
        if self.data is None:
            Display.show_message_box('Error', 'No data to save.')
            return

        try:
            with pd.ExcelWriter(self.selected_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                self.data.to_excel(writer, sheet_name=self.sheet_name, index=False)
            Display.show_message_box('Success', 'Data saved successfully.')
        except Exception as e:
            Display.show_message_box('Error', str(e))

    def get_data(self):
        return self.data

    def set_data(self, data):
        self.data = data

    def get_data_by_row(self, row_idx):
        if self.data is not None and 0 <= row_idx < len(self.data):
            return self.data.iloc[row_idx]
        else:
            return None
    def test_api_connection(self):
        try:
            self.con = OpenAIConnection()
            self.openAIclient2 = self.con.get_connection()               
            workbook = openpyxl.load_workbook(self.selected_file)
            sheet = workbook[self.sheet_name]
            cell_range = self.range_input.text()
            cells = sheet[cell_range]
            self.inputText = ""
            self.outputText = ""
            self.rollingText = ""
            for row in cells:
                for cell in row:
                    if cell.value is not None:
                        self.inputText = self.user_prompt.toPlainText() + "\n" + str(cell.value)
                        response = self.openAIclient2.chat.completions.create(
                            model=self.model.text(),
                            messages=[
                                {
                                    "role": "system",
                                    "content": self.system_prompt.toPlainText()
                                },
                                {
                                    "role": "user",
                                    "content": self.inputText
                                }
                            ],
                            max_tokens=self.max_tokens_slider.value(),
                            top_p=1
                        )
                        self.outputText = response.choices[0].message.content
                        self.rollingText = self.rollingText + "\n" + self.outputText
                        cell.value = self.outputText                        
            print("in config dialog")
         

            Display.show_message_box("SuccessExtend", "API connection successful Extend!\nResponse: " + response.choices[0].message.content)
            self.responseAPItext.setText(response.choices[0].message.content)
            save_path, _ = QFileDialog.getSaveFileName(self, "Save File", "", "Excel Files (*.xlsx)")
            if not save_path:
                Display.show_message_box( "Error", "No save location selected!")
                return
            workbook.save(save_path)
            Display.show_message_box( "Success", f"File saved as {save_path}")   
        except Exception as e:
            print("Error", f"API connection failed!\nError: {str(e)}")
            Display.show_message_box( "Error", f"API connection failed!\nError: {str(e)}")

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    node = ExcelProcess_Node()
    node.show()
    sys.exit(app.exec())
