from PySide6 import QtWidgets
from PySide6.QtWidgets import QLabel, QTextEdit, QLineEdit, QPushButton, QComboBox, QFileDialog, QMessageBox, QVBoxLayout, QFormLayout
from core.node import Node
from core.configdialog import ConfigDialog
from core.common import Node_Status
from openpyxl.utils import range_boundaries, get_column_letter
from utils.display import Display
from utils.llmconnection import LLMConnection
from utils.datamodels import SelectedLLM
import utils.themecolors as colors
import openpyxl
import pandas as pd
import json
import sys
import traceback


class ExcelAdvancedProcess_Node(Node):
    def __init__(self,node_config = {}):
        super().__init__()
        self.title_text = "Excel Interface"
        self.type_text = "Work with Excel"
        self.set_color(colors.get_color_rgb("output"))
        self.pin_output = self.add_pin(name="value", is_output=True)
        self.pin_A = self.add_pin(name="input A", is_output=False)
        self.build(node_config)

        self.config = {
            "user_prompt": "please write a detailed report on the below material",
            "system_prompt": "You are a security risk professional",
            "output":"",
            "additional_input":"",
            "max_tokens": 1024,
            "id": "",
            "object": "",
            "temperature": ""
        }

        # Create a single instance of the ExtendedConfigDialog
        
        
    def init_widget(self, node_config):
        super().init_widget()        
        self.config = node_config        
        if node_config == {}:
            self.config = {
                "user_prompt": "please write a detailed report on the below material",
                "system_prompt": "You are a security risk professional",
                "output":"",
                "additional_input":"",
                "max_tokens": 1024,
                "id": "",
                "object": "",
                "temperature": ""
            }  
        else: 
            self.config = node_config

        self.widget = QtWidgets.QWidget()
        self.widget.setFixedWidth(200)
        layout = QtWidgets.QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        self.textbox = QTextEdit()
        self.textbox.setFixedHeight(50)
        self.responsetext = QTextEdit()
        self.responsetext.setFixedHeight(50)
        self.responsetext.setStyleSheet("""
        QTextEdit{
        background: rgb(50, 50, 50); /*background color */
        }
        """)
        self.responsetext.setReadOnly(True)

        self.btn = QtWidgets.QPushButton("Refresh")
        self.btn.clicked.connect(self.btn_refresh)

        self.config_btn = QtWidgets.QPushButton("Configure")
        self.config_btn.clicked.connect(self.show_configuration)

        layout.addWidget(self.textbox)
        layout.addWidget(self.responsetext)
        layout.addWidget(self.btn)
        layout.addWidget(self.config_btn)
        self.widget.setLayout(layout)

        proxy = QtWidgets.QGraphicsProxyWidget()
        proxy.setWidget(self.widget)
        proxy.setParentItem(self)

    def show_configuration(self):
        self.btn_refresh()
        if self.pin_A:
            self.config['additional_input'] = str(self.pin_A.connected_pin.get_data())
            print("[Excel Process] added data:",self.config['additional_input'])
        main_window = QtWidgets.QMainWindow()
        self.dialog = ExtendedConfigDialog(json.dumps(self.config), main_window)            
        if self.dialog.exec():
            self.config = json.loads(self.dialog.get_configuration())
            self.responsetext.setText(self.config["output"])
            self.textbox.setText(self.dialog.complete_prompt.toPlainText())
            self.pin_output.set_data(self.responsetext.toPlainText())
        print("Pin output data:", self.pin_output.get_data())
        if self.pin_output.get_data() != "":
            self.status = Node_Status.CLEAN

    def btn_refresh(self):
        print("Refreshing data")
        if self.pin_A:
            self.pin_A.set_data(self.pin_A.connected_pin.get_data())
            print("pin A set:",self.pin_A.connected_pin.get_data())
            self.responsetext.setText(str(self.pin_A.connected_pin.get_data()))
            self.config['additional_input'] = str(self.pin_A.connected_pin.get_data())
            print(self.config)
            #if self.dialog.update_input(str(self.pin_A.connected_pin.get_data())):
            #print(self.config)


class ExtendedConfigDialog(ConfigDialog):

    
    def __init__(self, config_json, parent=None):
        super(ExtendedConfigDialog, self).__init__(config_json, parent)
        self.min_col = 1
        self.min_row = 1 
        self.max_col = 1
        self.max_row = 1


        self.file_label = QLabel('Selected file:')
        self.sheet_label = QLabel('Sheet name:')

        self.range_input = QLineEdit(self)
        self.range_input.setPlaceholderText('Enter Excel range (e.g., A1:B10)')

        self.sheet_combobox = QComboBox(self)
        self.sheet_combobox.setEnabled(False)

        self.load_button = QPushButton('Load Excel File', self)
        self.load_button.clicked.connect(self.load_file)

        self.get_data_button = QPushButton('Check Data', self)
        self.get_data_button.clicked.connect(self.get_range)
        self.get_data_button.setEnabled(False)

        self.excel_combobox = QComboBox(self)
        self.excel_combobox.addItem("Update all selected data")
        self.excel_combobox.addItem("Update based on row and column")
        self.excel_combobox.addItem("Update only highlighted data")
        self.excel_combobox.setEnabled(False)

        self.excel_combobox.currentIndexChanged.connect(self.excel_combobox_updated)

        self.top_layout = QFormLayout()
        self.top_layout.addWidget(self.load_button)
        self.top_layout.addWidget(self.sheet_label)
        self.top_layout.addWidget(self.file_label)
        self.top_layout.addWidget(self.sheet_combobox)
        self.top_layout.addWidget(self.range_input)
        self.top_layout.addWidget(self.get_data_button)
        self.top_layout.addWidget(self.excel_combobox)
        
        self.main_layout.addLayout(self.top_layout)
        self.setmainlayout()
        self.setLayout(self.main_layout)

        # Initialize message box attributes
        self.message_box = None

    def get_configuration(self):
        config = json.loads(super().get_configuration())
        return json.dumps(config)

    def excel_combobox_updated(index):
        pass
    

    def updatecompleteprompt(self):
            if self.excel_combobox.currentIndex() == 0:
                tempString = self.user_prompt.toPlainText() + " [Excel Cell]" + self.post_prompt.toPlainText() + "\n" + self.system_prompt.toPlainText()  \
                    + "\n" +self.additional_input
            elif self.excel_combobox.currentIndex() == 1:
                tempString = self.user_prompt.toPlainText() + "[Row Header] and [Column Header]" + self.post_prompt.toPlainText() + "\n"  \
                    + self.system_prompt.toPlainText() + "\n" + self.additional_input        
            else:
                tempString = "NOT IMPLEMENTED"
            self.complete_prompt.setText(tempString)    
    
    def load_file(self):
        file_dialog = QFileDialog()
        file_dialog.setFileMode(QFileDialog.ExistingFile)
        file_dialog.setNameFilter("Excel Files (*.xlsx *.xls)")

        if file_dialog.exec():
            self.selected_file = file_dialog.selectedFiles()[0]
            self.file_label.setText(f'Selected file: {self.selected_file}')

            xls = pd.ExcelFile(self.selected_file)
            self.sheet_combobox.clear()
            self.sheet_combobox.addItems(xls.sheet_names)
            if xls.sheet_names:
                self.sheet_combobox.setCurrentIndex(0)
                self.sheet_combobox.setEnabled(True)
                self.get_data_button.setEnabled(True)
                self.excel_combobox.setEnabled(True)
            
    def get_range(self):
        if not self.selected_file:
            Display.show_message_box('Error', 'No file selected.')
            Display.show_message_box("error","sdfs")
            return

        range_str = self.range_input.text()
        self.sheet_name = self.sheet_combobox.currentText()

        try:
            df = pd.read_excel(self.selected_file, sheet_name=self.sheet_name, engine='openpyxl',header=None)
            if range_str:
                df = self.get_data_in_range(df, range_str)
            self.data = df
            print(self.data)
            Display.show_message_box('Success', f'Selected range:\n{df}')
        except Exception as e:
            self.show_message_box('Error', str(e))

    def get_data_in_range(self, df, range_str):
        self.min_col, self.min_row, self.max_col, self.max_row = range_boundaries(range_str)
        return df.iloc[(self.min_row-1):(self.max_row), (self.min_col-1):(self.max_col)]

    def save_data(self):
        if self.data is None:
            Display.show_message_box('Error', 'No data to save.')
            return
        try:
            with pd.ExcelWriter(self.selected_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                self.data.to_excel(writer, sheet_name=self.sheet_name, index=False)
            Display.show_message_box('Success', 'Data saved successfully.')
        except Exception as e:
            Display.show_message_box('Error', str(e))

    def get_data(self):
        return self.data

    def set_data(self, data):
        self.data = data

    def get_data_by_row(self, row_idx):
        if self.data is not None and 0 <= row_idx < len(self.data):
            return self.data.iloc[row_idx]
        else:
            return None
        
    def execute_llm_connection(self):
        if self.selected_file == "":
            Display.show_message_box("Error", "No file selected")
            return None
        try:
            sLLM = SelectedLLM()
            workbook = openpyxl.load_workbook(self.selected_file)
            sheet = workbook[self.sheet_name]
            cell_range = self.range_input.text()
            cells = sheet[cell_range]
            connection = LLMConnection()
            self.selected_company = sLLM.selected_company
            inputText = ""
            connection.initiate_assistant("Excel Iterator",self.system_prompt.toPlainText())
            assistant_output = connection.call_assistant("Here is data for reference:\n\n" + self.additional_input, self.max_tokens_slider.value())            
            inputText = ""
            self.outputText = ""
            self.rollingText = ""
            #for condition where we want to answer all non-empty cells
            #self.excel_combobox.currentData
            if self.excel_combobox.currentIndex() == 0:
                print("Question mode")
                for row in cells:
                    for cell in row:
                        if cell.value is not None:
                            inputText = self.user_prompt.toPlainText() + str(cell.value) + self.post_prompt.toPlainText()
                            assistant_output = connection.call_assistant(inputText, self.max_tokens_slider.value(),embeddings=self.use_reference_data_checkbox.isChecked())            
                            #response = connection.call_prompt(inputText, self.system_prompt.toPlainText(),self.model.text(),self.max_tokens_slider.value())
                            self.outputText = assistant_output
                            print("---------------------------------------------------")
                            print(self.outputText)
                            self.rollingText = assistant_output + "\n" + self.rollingText
                            cell.value = assistant_output
                            self.responseAPItext.setText(self.rollingText)
            #for condition where we want to answer based on the row and column header
            else:
                print("Matrix mode")
                start_row = cells[0][0].row
                start_col = cells[0][0].column
                for row in cells:
                    row_number = row[0].row  # Get the row number (assuming first cell in row has it)
                    for cell in row:
                        if cell.value is None:
                            row_header = sheet.cell(row=row_number, column=start_col).value 
                            col_header = sheet.cell(row=start_row, column=cell.column).value                                               
                            inputText = self.user_prompt.toPlainText() + str(f"[{row_header}] and [{col_header}]") + \
                                        self.post_prompt.toPlainText() +  "\n" 
                            assistant_output = connection.call_assistant(inputText, self.max_tokens_slider.value(),embeddings=self.use_reference_data_checkbox.isChecked())            
                            self.outputText = assistant_output
                            print("---------------------------------------------------")
                            print(self.outputText)
                            self.rollingText = assistant_output + "\n" + self.rollingText
                            cell.value = assistant_output
                            self.responseAPItext.setText(self.rollingText)                                        
            print("in config dialog")

            Display.show_message_box("SuccessExtend", "API connection successful Extend!\nResponse: " + self.rollingText)
            self.responseAPItext.setText(self.rollingText)
            save_path, _ = QFileDialog.getSaveFileName(self, "Save File", "", "Excel Files (*.xlsx)")
            if not save_path:
                Display.show_message_box("Error", "No save location selected!")
                return
            workbook.save(save_path)
            Display.show_message_box("Success", f"File saved as {save_path}")
        except Exception as e:
            print("Error", f"API connection failed!\nError: {str(e)}")
            print(traceback.format_exc())
            Display.show_message_box("Error", f"API connection failed!\nError: {str(e)}")


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    node = ExcelAdvancedProcess_Node()
    node.show()
    sys.exit(app.exec())
